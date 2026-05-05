"""Data manager — parse, shape-detect, and register uploaded files."""

import csv
import json
import sqlite3
import xml.etree.ElementTree as ET
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from xlrd import open_workbook as open_xls_workbook

UPLOAD_DIR = Path("/app/upload")

_JSON_EXTS = {".json", ".jsonl", ".ndjson"}
_CSV_EXTS = {".csv", ".tsv", ".psv"}
_EXCEL_EXTS = {".xlsx", ".xls"}
_SQLITE_EXTS = {".sqlite", ".sqlite3", ".db"}
_SUPPORTED_EXTS = _JSON_EXTS | _CSV_EXTS | _EXCEL_EXTS | _SQLITE_EXTS | {".xml"}


def _decode_text(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("Unable to decode file text")


# ── JSON / JSONL parsing ──────────────────────────────────────

def _parse_json_lines(raw: str) -> list[Any]:
    items: list[Any] = []
    for idx, line in enumerate(raw.splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        try:
            items.append(json.loads(line))
        except json.JSONDecodeError as exc:
            raise ValueError(f"Invalid JSONL at line {idx}: {exc.msg}") from exc
    return items


def _parse_uploaded_json(filename: str, raw: str) -> Any:
    suffix = Path(filename).suffix.lower()
    if suffix in {".jsonl", ".ndjson"}:
        return _parse_json_lines(raw)

    try:
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        if "Extra data" in str(exc):
            return _parse_json_lines(raw)
        raise


# ── CSV / TSV / PSV parsing ───────────────────────────────────

def _csv_delimiter_for_suffix(suffix: str) -> str:
    if suffix == ".tsv":
        return "\t"
    if suffix == ".psv":
        return "|"
    return ","


def _parse_csv_content(filename: str, text: str) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    fallback_delim = _csv_delimiter_for_suffix(suffix)

    sample = "\n".join(text.splitlines()[:10])
    delimiter = fallback_delim
    try:
        sniffed = csv.Sniffer().sniff(sample, delimiters=[",", "\t", "|", ";"])
        delimiter = sniffed.delimiter
    except Exception:
        pass

    reader = csv.DictReader(StringIO(text), delimiter=delimiter)
    rows: list[dict[str, Any]] = []
    for row in reader:
        rec: dict[str, Any] = {}
        for key, value in row.items():
            clean_key = str(key or "").strip()
            if clean_key:
                rec[clean_key] = value
        if rec:
            rows.append(rec)
    return rows


# ── XML parsing ────────────────────────────────────────────────

def _xml_local(tag: str) -> str:
    if "}" in tag:
        return tag.rsplit("}", 1)[1]
    return tag


def _xml_to_object(elem: ET.Element) -> Any:
    attrs = {f"@{_xml_local(k)}": v for k, v in elem.attrib.items()}
    children = list(elem)
    text = (elem.text or "").strip()

    if not children and not attrs:
        return text

    out: dict[str, Any] = {}
    out.update(attrs)

    grouped: dict[str, list[Any]] = {}
    for child in children:
        key = _xml_local(child.tag)
        grouped.setdefault(key, []).append(_xml_to_object(child))

    for key, vals in grouped.items():
        out[key] = vals[0] if len(vals) == 1 else vals

    if text:
        out["#text"] = text
    return out


def _parse_xml_content(text: str) -> Any:
    root = ET.fromstring(text)
    root_obj = _xml_to_object(root)

    if isinstance(root_obj, dict):
        # Prefer list-of-records when present under root.
        for value in root_obj.values():
            if isinstance(value, list) and value and all(isinstance(v, dict) for v in value):
                return value
        return [root_obj]
    if isinstance(root_obj, list):
        return root_obj
    return [{_xml_local(root.tag): root_obj}]


# ── Excel parsing ──────────────────────────────────────────────

def _clean_header(value: Any, idx: int) -> str:
    text = str(value).strip() if value is not None else ""
    return text or f"column_{idx + 1}"


def _parse_xlsx(raw: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_clean_header(v, i) for i, v in enumerate(rows[0])]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        rec: dict[str, Any] = {}
        for i, value in enumerate(row):
            if i >= len(headers):
                break
            rec[headers[i]] = value
        if any(v is not None and str(v) != "" for v in rec.values()):
            out.append(rec)
    return out


def _parse_xls(raw: bytes) -> list[dict[str, Any]]:
    wb = open_xls_workbook(file_contents=raw)
    if wb.nsheets < 1:
        return []

    sh = wb.sheet_by_index(0)
    if sh.nrows < 1:
        return []

    headers = [_clean_header(sh.cell_value(0, i), i) for i in range(sh.ncols)]
    out: list[dict[str, Any]] = []
    for r in range(1, sh.nrows):
        rec = {headers[c]: sh.cell_value(r, c) for c in range(sh.ncols)}
        if any(v not in (None, "") for v in rec.values()):
            out.append(rec)
    return out


# ── SQLite helpers ─────────────────────────────────────────────

def _sqlite_tables(path: Path) -> list[str]:
    with sqlite3.connect(str(path)) as conn:
        rows = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
        ).fetchall()
    return [str(r[0]) for r in rows]


def _parse_uploaded_data(filename: str, raw: bytes) -> Any:
    suffix = Path(filename).suffix.lower()
    if suffix not in _SUPPORTED_EXTS:
        raise ValueError(f"Unsupported extension: {suffix}")

    if suffix in _JSON_EXTS:
        return _parse_uploaded_json(filename, _decode_text(raw))
    if suffix in _CSV_EXTS:
        return _parse_csv_content(filename, _decode_text(raw))
    if suffix == ".xml":
        return _parse_xml_content(_decode_text(raw))
    if suffix == ".xlsx":
        return _parse_xlsx(raw)
    if suffix == ".xls":
        return _parse_xls(raw)
    if suffix in _SQLITE_EXTS:
        return {"_sqlite": True}

    raise ValueError(f"Unsupported extension: {suffix}")


# ── Tabular normalization ──────────────────────────────────────

def _flatten_record(record: dict, prefix: str = "") -> dict:
    out: dict[str, Any] = {}
    for key, value in record.items():
        name = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            out.update(_flatten_record(value, name))
        elif isinstance(value, list):
            out[name] = json.dumps(value, ensure_ascii=False)
        else:
            out[name] = value
    return out


def _normalize_for_tabular(data: Any) -> tuple[Any, bool]:
    if isinstance(data, list) and data and all(isinstance(r, dict) for r in data):
        flattened = [_flatten_record(r) for r in data]
        transformed = any(any("." in k for k in row.keys()) for row in flattened)
        return flattened, transformed

    if isinstance(data, dict):
        flattened = _flatten_record(data)
        transformed = any("." in k for k in flattened.keys())
        return [flattened], transformed

    return data, False


# ── Shape detection ────────────────────────────────────────────

def detect_shape(data: Any) -> str:
    if isinstance(data, list):
        if not data:
            return "empty_array"
        first = data[0]
        if isinstance(first, dict):
            return "array_of_objects"
        if isinstance(first, list):
            return "array_of_arrays"
        return "primitive_array"
    if isinstance(data, dict):
        has_nested = any(isinstance(v, (dict, list)) for v in data.values())
        return "nested_object" if has_nested else "flat_object"
    return "primitive"


# ── Tree builder ───────────────────────────────────────────────

def _build_tree(data: Any, max_depth: int = 10, depth: int = 0) -> Any:
    if depth >= max_depth:
        return "..."
    if isinstance(data, dict):
        items = list(data.items())
        truncated = len(items) > 100
        out = {k: _build_tree(v, max_depth, depth + 1) for k, v in items[:100]}
        if truncated:
            out["__truncated__"] = f"({len(items) - 100} more keys)"
        return out
    if isinstance(data, list):
        if len(data) > 200:
            return (
                [_build_tree(it, max_depth, depth + 1) for it in data[:200]]
                + [f"... ({len(data) - 200} more items)"]
            )
        return [_build_tree(it, max_depth, depth + 1) for it in data]
    return data


# ── Manual schema fallback ─────────────────────────────────────

def _manual_schema(data: Any, shape: str) -> list[dict]:
    if shape == "array_of_objects" and data:
        return [{"name": k, "type": type(v).__name__} for k, v in data[0].items()]
    if shape in ("flat_object", "nested_object") and isinstance(data, dict):
        return [{"name": k, "type": type(v).__name__} for k, v in data.items()]
    return [{"name": "value", "type": "any"}]


def _register_hint_for_suffix(suffix: str) -> str:
    if suffix in _CSV_EXTS:
        return "csv"
    if suffix in _SQLITE_EXTS:
        return "sqlite"
    return "json"


# ── Public API ────────────────────────────────────────────────

def load_and_register(filename: str) -> dict:
    """Parse a supported file and register it with DuckDB. Returns metadata."""
    from services.duck_engine import duck_engine

    path = UPLOAD_DIR / filename
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filename}")

    resolved = path.resolve()
    if not str(resolved).startswith(str(UPLOAD_DIR.resolve())):
        raise ValueError("Invalid file path (path traversal detected)")

    suffix = resolved.suffix.lower()
    size = resolved.stat().st_size

    raw = resolved.read_bytes()
    data = _parse_uploaded_data(filename, raw)
    shape = "sqlite_database" if suffix in _SQLITE_EXTS else detect_shape(data)

    normalized_data, was_transformed = _normalize_for_tabular(data)

    register_path = resolved
    register_hint = _register_hint_for_suffix(suffix)
    register_options: dict[str, Any] = {}

    if suffix in _SQLITE_EXTS:
        tables = _sqlite_tables(resolved)
        if not tables:
            raise ValueError("SQLite file has no user tables")
        register_options["table"] = tables[0]
    elif suffix in _CSV_EXTS:
        register_options["delimiter"] = _csv_delimiter_for_suffix(suffix)

    # If parser produced in-memory rows, write normalized JSON for registration.
    if suffix in _EXCEL_EXTS or suffix == ".xml" or (register_hint == "json" and was_transformed):
        normalized_dir = UPLOAD_DIR / ".normalized"
        normalized_dir.mkdir(parents=True, exist_ok=True)
        normalized_path = normalized_dir / f"{Path(filename).stem}.normalized.json"
        normalized_path.write_text(
            json.dumps(normalized_data, ensure_ascii=False),
            encoding="utf-8",
        )
        register_path = normalized_path
        register_hint = "json"

    schema: list[dict] = []
    row_count = 0
    view_name = duck_engine.view_name(filename)
    error: str | None = None

    try:
        result = duck_engine.register(
            filename,
            register_path,
            format_hint=register_hint,
            options=register_options,
        )
        schema = result["schema"]
        row_count = duck_engine.row_count(view_name)
    except Exception as exc:
        error = str(exc)
        schema = _manual_schema(data, shape)
        if isinstance(data, list):
            row_count = len(data)
        elif isinstance(data, dict):
            row_count = 1

    meta = {
        "filename": filename,
        "view_name": view_name,
        "shape": shape,
        "schema": schema,
        "row_count": row_count,
        "size": size,
    }
    if error:
        meta["warning"] = f"DuckDB registration issue: {error}"
    elif suffix in _SQLITE_EXTS and register_options.get("table"):
        meta["warning"] = f"Loaded first SQLite table: {register_options['table']}"
    elif was_transformed:
        meta["warning"] = "Nested fields were flattened for table view (dot notation columns)."
    return meta


def get_tree(filename: str) -> Any:
    """Read a supported file from disk and return a truncated tree structure."""
    path = UPLOAD_DIR / filename
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filename}")

    resolved = path.resolve()
    if not str(resolved).startswith(str(UPLOAD_DIR.resolve())):
        raise ValueError("Invalid file path")

    suffix = resolved.suffix.lower()
    if suffix in _SQLITE_EXTS:
        data = {"type": "sqlite_database", "tables": _sqlite_tables(resolved)}
    else:
        data = _parse_uploaded_data(filename, resolved.read_bytes())
    return _build_tree(data)


def list_files() -> list[dict]:
    """List all supported files in the upload directory."""
    if not UPLOAD_DIR.exists():
        return []

    result = []
    for f in sorted(UPLOAD_DIR.iterdir()):
        if f.is_file() and f.suffix.lower() in _SUPPORTED_EXTS and not f.name.startswith("."):
            result.append({"filename": f.name, "size": f.stat().st_size})
    return result
